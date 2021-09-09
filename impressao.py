from selenium import webdriver
import pyautogui
from time import sleep
import openpyxl


class ChromeAuto():
    def __init__(self):
        self.driver_path = 'chromedriver'
        self.options = webdriver.ChromeOptions()
        self.chrome = webdriver.Chrome(
            self.driver_path,
            options=self.options
        )

    def acessa(self, site):
        self.chrome.get(site)

    def clica_label_print(self):
        try:
            btn_next_page = self.chrome.find_element_by_id('TableNextPage')
            sleep(1)
            btn_next_page.click()
            sleep(1)
            pyautogui.keyDown('f8')
            pyautogui.keyUp('f8')
            sleep(1)
            pyautogui.keyDown('f4')
            pyautogui.keyUp('f4')
            sleep(1)
        except Exception as e:
            print('Erro ao clicar em Label Print: ', e)

    def faz_login(self):
        try:
            input_login = self.chrome.find_element_by_id('sap-user')
            input_password = self.chrome.find_element_by_id('sap-password')
            btn_login = self.chrome.find_element_by_class_name('MobileLoginStdBtn')

            input_login.send_keys('ID')
            input_password.send_keys('Password')
            sleep(1)
            btn_login.click()

        except Exception as e:
            print('Erro ao fazer login: ',e)

    def faz_logout(self):
        try:
            btn_prev = self.chrome.find_element_by_id('prev[1]')
            btn_prev.click()
            sleep(1)
            btn_prev = self.chrome.find_element_by_id('prev[1]')
            btn_prev.click()
            sleep(1)
            btn_logout = self.chrome.find_element_by_id('pb_logout[1]')
            btn_logout.click()
            sleep(1)
            print('Etiquetas impressas.')

        except Exception as e:
            print('Erro ao fazer logout: ', e)

    def imprimir_etiqueta(self, material, batch, quantity):
        try:
            input_material = self.chrome.find_element_by_id('gwa_user_input-matnr[1]')
            input_material.send_keys(str(material))
            sleep(1)
            pyautogui.keyDown('enter')
            pyautogui.keyUp('enter')
            sleep(1)
            input_batch = self.chrome.find_element_by_id('gwa_user_input-charg[1]')
            sleep(1)
            input_batch.send_keys(str(batch))
            pyautogui.keyDown('enter')
            pyautogui.keyUp('enter')
            sleep(1)
            input_quantity = self.chrome.find_element_by_id('gwa_user_input-gesme_d[1]')
            sleep(1)
            input_quantity.send_keys(str(quantity))
            pyautogui.keyDown('enter')
            pyautogui.keyUp('enter')
            sleep(1)
            btn_print = self.chrome.find_element_by_id('print[1]')
            btn_print.click()
            sleep(1)
        except Exception as e:
            print('Erro ao imprimir etiqueta: ', e)

    def sair(self):
        self.chrome.quit()


if __name__ == '__main__':
    chrome = ChromeAuto()
    chrome.acessa('http://siteimpressaodeetiqueta.com')
    chrome.faz_login()
    chrome.clica_label_print()
    sleep(1)
    amostras = openpyxl.load_workbook(r'Amostras.xlsm')
    planilha = amostras['Baixa de Materiais']
    for linha in planilha.rows:
        material = planilha.cell(linha, 4).value
        batch = planilha.cell(linha, 5).value
        quantity = planilha.cell(linha, 8).value
        chrome.imprimir_etiqueta(material, batch, quantity)
    amostras.close()
    sleep(1)
    chrome.faz_logout()
    chrome.sair()